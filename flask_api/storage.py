
import os
import json
from openpyxl import Workbook, load_workbook
from datetime import datetime

COMMENTS_FILE = "comments_log_ai.xlsx"
KEYWORDS_FILE = os.path.join(os.path.dirname(__file__), "ai_keywords.json")

# Загрузка ключевых слов
with open(KEYWORDS_FILE, encoding="utf-8") as f:
    CATEGORY_KEYWORDS = json.load(f)

def classify_comment(comment):
    comment = comment.lower()
    for category, words in CATEGORY_KEYWORDS.items():
        if any(kw in comment for kw in words):
            return category
    return "нейтральный"

def init_excel():
    if not os.path.exists(COMMENTS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "AI_Comments"
        ws.append(["timestamp", "author", "comment", "category", "likes", "post_id", "parent_id"])
        wb.save(COMMENTS_FILE)

def save_comment_ai(comment, author=None, post_id=None, likes=0, parent_id=""):
    init_excel()
    wb = load_workbook(COMMENTS_FILE)
    ws = wb.active

    category = classify_comment(comment)
    ws.append([
        datetime.utcnow().isoformat(),
        author or "anonymous",
        comment,
        category,
        likes,
        post_id or "unknown",
        parent_id
    ])
    wb.save(COMMENTS_FILE)

    return {
        "timestamp": datetime.utcnow().isoformat(),
        "author": author,
        "comment": comment,
        "category": category,
        "likes": likes,
        "post_id": post_id,
        "parent_id": parent_id
    }

def load_comments_ai():
    if not os.path.exists(COMMENTS_FILE):
        return []

    wb = load_workbook(COMMENTS_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = [dict(zip(headers, row)) for row in rows[1:]]
    return data
